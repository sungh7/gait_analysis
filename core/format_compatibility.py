#!/usr/bin/env python3
"""
MediaPipe 보행분석 결과 형식 호환성 모듈
- 병원 Excel 형식과 완전 호환
- 전통적 보행분석 시스템 결과 형식 지원
- S1_*.xlsx 구조 완벽 재현
- 다양한 출력 형식 지원

Author: AI Assistant
Date: 2025-09-15
"""

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from pathlib import Path
import json
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import LineChart, Reference
import warnings
warnings.filterwarnings('ignore')

class FormatCompatibility:
    """형식 호환성 모듈"""

    def __init__(self):
        self.conversion_results = {}

        # 표준 Excel 형식 템플릿
        self.excel_template = {
            'Main_Data': ['Subject_ID', 'Age', 'Height', 'Weight', 'Gender', 'Analysis_Date'],
            'Discrete_Parameters': [
                'Cadence', 'Stride_Length', 'Stride_Time', 'Step_Time', 'Walking_Speed',
                'Stance_Phase_Percent', 'Swing_Phase_Percent', 'Double_Support_Percent'
            ],
            'Joint_Angles_101': [
                'Gait_Cycle_Percent', 'Hip_Flexion', 'Knee_Flexion', 'Ankle_Dorsiflexion'
            ],
            'Temporal_Spatial': [
                'Parameter', 'Mean', 'Std', 'Min', 'Max', 'Unit'
            ]
        }

        print("✅ 형식 호환성 모듈 초기화 완료")

    def convert_mediapipe_to_excel(self, mp_results, output_path, subject_info=None):
        """MediaPipe 결과를 표준 Excel 형식으로 변환"""
        print(f"📊 Excel 형식 변환 시작: {output_path}")

        # 기본 피험자 정보
        if subject_info is None:
            subject_info = {
                'Subject_ID': mp_results.get('subject_id', 'Unknown'),
                'Age': 'N/A',
                'Height': 'N/A',
                'Weight': 'N/A',
                'Gender': 'N/A',
                'Analysis_Date': datetime.now().strftime('%Y-%m-%d')
            }

        # Excel 워크북 생성
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:

            # 1. Main_Data 시트
            self._create_main_data_sheet(subject_info, writer)

            # 2. Discrete_Parameters 시트
            self._create_discrete_parameters_sheet(mp_results, writer)

            # 3. Joint_Angles_101 시트
            self._create_joint_angles_sheet(mp_results, writer)

            # 4. Temporal_Spatial 시트
            self._create_temporal_spatial_sheet(mp_results, writer)

            # 5. Summary 시트
            self._create_summary_sheet(mp_results, writer)

        # Excel 스타일링 적용
        self._apply_excel_styling(output_path)

        print(f"✅ Excel 변환 완료: {output_path}")
        return output_path

    def _create_main_data_sheet(self, subject_info, writer):
        """메인 데이터 시트 생성"""
        main_data = pd.DataFrame([subject_info])
        main_data.to_excel(writer, sheet_name='Main_Data', index=False)

    def _create_discrete_parameters_sheet(self, mp_results, writer):
        """이산 매개변수 시트 생성"""
        ts_data = mp_results.get('temporal_spatial', {})

        # 매개변수 매핑
        param_mapping = {
            'Cadence': ts_data.get('cadence', 0),
            'Stride_Length': ts_data.get('stride_length_mean', 0),
            'Stride_Time': ts_data.get('stride_time_mean', 0),
            'Step_Time': ts_data.get('step_time_mean', 0),
            'Walking_Speed': ts_data.get('walking_speed_mean', 0),
            'Stance_Phase_Percent': ts_data.get('stance_phase_mean', 60),
            'Swing_Phase_Percent': ts_data.get('swing_phase_mean', 40),
            'Double_Support_Percent': np.mean(ts_data.get('double_support_percent', [0]))
        }

        # 리스트 데이터 처리 (여러 보행주기가 있는 경우)
        discrete_data = []

        # 최대 길이 찾기
        max_cycles = 1
        for key in ['stride_length_list', 'stride_time_list', 'step_time_list', 'walking_speed_list']:
            if key in ts_data and ts_data[key]:
                max_cycles = max(max_cycles, len(ts_data[key]))

        # 각 보행주기별 데이터 생성
        for i in range(max_cycles):
            cycle_data = {}
            for param, default_value in param_mapping.items():
                # 리스트 데이터가 있으면 해당 인덱스 사용, 없으면 기본값
                list_key = self._get_list_key_for_param(param)
                if list_key and list_key in ts_data and ts_data[list_key] and i < len(ts_data[list_key]):
                    cycle_data[param] = ts_data[list_key][i]
                else:
                    cycle_data[param] = default_value

            discrete_data.append(cycle_data)

        discrete_df = pd.DataFrame(discrete_data)
        discrete_df.to_excel(writer, sheet_name='Discrete_Parameters', index=False)

    def _get_list_key_for_param(self, param):
        """매개변수에 대응하는 리스트 키 반환"""
        mapping = {
            'Stride_Length': 'stride_length_list',
            'Stride_Time': 'stride_time_list',
            'Step_Time': 'step_time_list',
            'Walking_Speed': 'walking_speed_list',
            'Stance_Phase_Percent': 'stance_phase_percent',
            'Swing_Phase_Percent': 'swing_phase_percent'
        }
        return mapping.get(param)

    def _create_joint_angles_sheet(self, mp_results, writer):
        """관절각도 101포인트 시트 생성"""
        joint_angles_101 = mp_results.get('joint_angles_101', {})

        # 101포인트 데이터 구성
        data = {
            'Gait_Cycle_Percent': list(range(0, 101)),
            'Hip_Flexion': joint_angles_101.get('hip_flexion_extension', [0] * 101),
            'Knee_Flexion': joint_angles_101.get('knee_flexion_extension', [0] * 101),
            'Ankle_Dorsiflexion': joint_angles_101.get('ankle_dorsi_plantarflexion', [0] * 101)
        }

        # 데이터 길이 보정 (101포인트 맞추기)
        for key, values in data.items():
            if key != 'Gait_Cycle_Percent':
                if len(values) < 101:
                    # 부족한 경우 보간
                    if len(values) > 1:
                        x_old = np.linspace(0, 100, len(values))
                        x_new = np.linspace(0, 100, 101)
                        from scipy.interpolate import interp1d
                        interp_func = interp1d(x_old, values, kind='linear', fill_value='extrapolate')
                        data[key] = interp_func(x_new).tolist()
                    else:
                        data[key] = [0] * 101
                elif len(values) > 101:
                    # 초과하는 경우 자르기
                    data[key] = values[:101]

        joint_df = pd.DataFrame(data)
        joint_df.to_excel(writer, sheet_name='Joint_Angles_101', index=False)

    def _create_temporal_spatial_sheet(self, mp_results, writer):
        """시공간 매개변수 요약 시트 생성"""
        ts_data = mp_results.get('temporal_spatial', {})

        # 요약 통계 생성
        summary_data = []

        parameters = [
            ('Cadence', 'cadence', 'cadence_list', 'steps/min'),
            ('Stride Length', 'stride_length_mean', 'stride_length_list', 'm'),
            ('Stride Time', 'stride_time_mean', 'stride_time_list', 's'),
            ('Step Time', 'step_time_mean', 'step_time_list', 's'),
            ('Walking Speed', 'walking_speed_mean', 'walking_speed_list', 'm/s'),
            ('Stance Phase', 'stance_phase_mean', 'stance_phase_percent', '%'),
            ('Swing Phase', 'swing_phase_mean', 'swing_phase_percent', '%')
        ]

        for param_name, mean_key, list_key, unit in parameters:
            # 평균값
            mean_val = ts_data.get(mean_key, 0)

            # 리스트에서 통계 계산
            if list_key in ts_data and ts_data[list_key]:
                values = ts_data[list_key]
                std_val = np.std(values)
                min_val = np.min(values)
                max_val = np.max(values)
            else:
                std_val = 0
                min_val = mean_val
                max_val = mean_val

            summary_data.append({
                'Parameter': param_name,
                'Mean': f"{mean_val:.3f}",
                'Std': f"{std_val:.3f}",
                'Min': f"{min_val:.3f}",
                'Max': f"{max_val:.3f}",
                'Unit': unit
            })

        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='Temporal_Spatial', index=False)

    def _create_summary_sheet(self, mp_results, writer):
        """요약 시트 생성"""
        # 분석 정보
        analysis_info = {
            'Analysis Information': [
                f"Subject ID: {mp_results.get('subject_id', 'Unknown')}",
                f"Analysis Date: {mp_results.get('analysis_timestamp', datetime.now().isoformat())}",
                f"Pipeline Version: {mp_results.get('pipeline_version', '1.0')}",
                f"Video Duration: {mp_results.get('video_info', {}).get('duration', 0):.2f} seconds",
                f"Video FPS: {mp_results.get('video_info', {}).get('fps', 0):.1f}",
                f"Total Frames: {mp_results.get('video_info', {}).get('frame_count', 0)}"
            ]
        }

        # 주요 결과
        ts_data = mp_results.get('temporal_spatial', {})
        key_results = {
            'Key Results': [
                f"Cadence: {ts_data.get('cadence', 0):.1f} steps/min",
                f"Walking Speed: {ts_data.get('walking_speed_mean', 0):.3f} m/s",
                f"Stride Length: {ts_data.get('stride_length_mean', 0):.3f} m",
                f"Stride Time: {ts_data.get('stride_time_mean', 0):.3f} s",
                f"Stance Phase: {ts_data.get('stance_phase_mean', 60):.1f}%",
                f"Swing Phase: {ts_data.get('swing_phase_mean', 40):.1f}%"
            ]
        }

        # 데이터프레임 생성
        summary_data = []
        max_len = max(len(analysis_info['Analysis Information']), len(key_results['Key Results']))

        for i in range(max_len):
            row = {}
            if i < len(analysis_info['Analysis Information']):
                row['Analysis Information'] = analysis_info['Analysis Information'][i]
            else:
                row['Analysis Information'] = ''

            if i < len(key_results['Key Results']):
                row['Key Results'] = key_results['Key Results'][i]
            else:
                row['Key Results'] = ''

            summary_data.append(row)

        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='Summary', index=False)

    def _apply_excel_styling(self, excel_path):
        """Excel 파일에 스타일 적용"""
        try:
            wb = openpyxl.load_workbook(excel_path)

            # 헤더 스타일
            header_font = Font(bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center')

            # 데이터 스타일
            data_alignment = Alignment(horizontal='center', vertical='center')
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # 각 시트에 스타일 적용
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]

                # 헤더 행 스타일링
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = border

                # 데이터 행 스타일링
                for row in ws.iter_rows(min_row=2):
                    for cell in row:
                        cell.alignment = data_alignment
                        cell.border = border

                # 컬럼 너비 자동 조정
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter

                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass

                    adjusted_width = min(max_length + 2, 20)
                    ws.column_dimensions[column_letter].width = adjusted_width

            wb.save(excel_path)
            print("✅ Excel 스타일링 적용 완료")

        except Exception as e:
            print(f"⚠️ Excel 스타일링 중 오류: {e}")

    def convert_to_csv_format(self, mp_results, output_dir):
        """CSV 형식으로 변환"""
        output_path = Path(output_dir)
        output_path.mkdir(exist_ok=True, parents=True)

        # 각 시트를 별도 CSV로 저장
        subject_id = mp_results.get('subject_id', 'unknown')

        # 1. 이산 매개변수 CSV
        ts_data = mp_results.get('temporal_spatial', {})
        discrete_data = {
            'Parameter': ['Cadence', 'Stride_Length', 'Stride_Time', 'Walking_Speed', 'Stance_Phase'],
            'Value': [
                ts_data.get('cadence', 0),
                ts_data.get('stride_length_mean', 0),
                ts_data.get('stride_time_mean', 0),
                ts_data.get('walking_speed_mean', 0),
                ts_data.get('stance_phase_mean', 60)
            ],
            'Unit': ['steps/min', 'm', 's', 'm/s', '%']
        }

        discrete_df = pd.DataFrame(discrete_data)
        discrete_csv = output_path / f"{subject_id}_discrete_parameters.csv"
        discrete_df.to_csv(discrete_csv, index=False)

        # 2. 관절각도 CSV
        joint_angles_101 = mp_results.get('joint_angles_101', {})
        joint_data = {
            'Gait_Cycle_Percent': list(range(0, 101)),
            'Hip_Flexion': joint_angles_101.get('hip_flexion_extension', [0] * 101)[:101],
            'Knee_Flexion': joint_angles_101.get('knee_flexion_extension', [0] * 101)[:101],
            'Ankle_Flexion': joint_angles_101.get('ankle_dorsi_plantarflexion', [0] * 101)[:101]
        }

        joint_df = pd.DataFrame(joint_data)
        joint_csv = output_path / f"{subject_id}_joint_angles_101.csv"
        joint_df.to_csv(joint_csv, index=False)

        print(f"✅ CSV 변환 완료: {output_path}")
        return [discrete_csv, joint_csv]

    def convert_to_json_standard(self, mp_results, output_path):
        """표준화된 JSON 형식으로 변환"""

        # 표준 JSON 구조
        standard_json = {
            "metadata": {
                "subject_id": mp_results.get('subject_id', 'unknown'),
                "analysis_timestamp": mp_results.get('analysis_timestamp', datetime.now().isoformat()),
                "pipeline_version": mp_results.get('pipeline_version', '1.0'),
                "video_info": mp_results.get('video_info', {})
            },
            "temporal_spatial_parameters": {
                "cadence": {
                    "value": mp_results.get('temporal_spatial', {}).get('cadence', 0),
                    "unit": "steps/min"
                },
                "stride_length": {
                    "mean": mp_results.get('temporal_spatial', {}).get('stride_length_mean', 0),
                    "std": mp_results.get('temporal_spatial', {}).get('stride_length_std', 0),
                    "unit": "m"
                },
                "walking_speed": {
                    "mean": mp_results.get('temporal_spatial', {}).get('walking_speed_mean', 0),
                    "std": mp_results.get('temporal_spatial', {}).get('walking_speed_std', 0),
                    "unit": "m/s"
                },
                "stance_phase": {
                    "mean": mp_results.get('temporal_spatial', {}).get('stance_phase_mean', 60),
                    "std": mp_results.get('temporal_spatial', {}).get('stance_phase_std', 0),
                    "unit": "percent"
                }
            },
            "joint_angles_normalized": {
                "hip_flexion_extension": {
                    "data": mp_results.get('joint_angles_101', {}).get('hip_flexion_extension', []),
                    "unit": "degrees",
                    "normalization": "101_points_gait_cycle"
                },
                "knee_flexion_extension": {
                    "data": mp_results.get('joint_angles_101', {}).get('knee_flexion_extension', []),
                    "unit": "degrees",
                    "normalization": "101_points_gait_cycle"
                },
                "ankle_dorsi_plantarflexion": {
                    "data": mp_results.get('joint_angles_101', {}).get('ankle_dorsi_plantarflexion', []),
                    "unit": "degrees",
                    "normalization": "101_points_gait_cycle"
                }
            },
            "gait_events": mp_results.get('gait_events', {}),
            "quality_metrics": {
                "total_frames_analyzed": len(mp_results.get('joint_angles_raw', {}).get('timestamps', [])),
                "successful_pose_detections": len([t for t in mp_results.get('joint_angles_raw', {}).get('timestamps', []) if t is not None])
            }
        }

        # JSON 저장
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(standard_json, f, ensure_ascii=False, indent=2)

        print(f"✅ 표준 JSON 변환 완료: {output_path}")
        return output_path

    def create_comparison_report(self, mp_results, traditional_results, output_path):
        """비교 분석 리포트 생성"""

        # 비교 데이터 추출
        mp_ts = mp_results.get('temporal_spatial', {})

        # 간단한 비교 테이블 생성
        comparison_data = {
            'Parameter': ['Cadence', 'Stride Length', 'Walking Speed', 'Stance Phase'],
            'MediaPipe': [
                f"{mp_ts.get('cadence', 0):.1f}",
                f"{mp_ts.get('stride_length_mean', 0):.3f}",
                f"{mp_ts.get('walking_speed_mean', 0):.3f}",
                f"{mp_ts.get('stance_phase_mean', 60):.1f}"
            ],
            'Traditional': ['N/A', 'N/A', 'N/A', 'N/A'],  # 전통적 결과가 있다면 채움
            'Difference': ['N/A', 'N/A', 'N/A', 'N/A'],
            'Unit': ['steps/min', 'm', 'm/s', '%']
        }

        comparison_df = pd.DataFrame(comparison_data)

        # Excel로 저장
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            comparison_df.to_excel(writer, sheet_name='Comparison', index=False)

            # MediaPipe 상세 결과도 포함
            self._create_discrete_parameters_sheet(mp_results, writer)
            self._create_joint_angles_sheet(mp_results, writer)

        print(f"✅ 비교 리포트 생성 완료: {output_path}")
        return output_path

def main():
    """메인 함수 - 형식 변환 테스트"""
    print("🔄 MediaPipe 결과 형식 변환 시스템")

    # 형식 변환기 초기화
    converter = FormatCompatibility()

    # 예시 MediaPipe 결과 (실제 분석 결과로 대체)
    sample_results = {
        'subject_id': 'S001',
        'temporal_spatial': {
            'cadence': 115.5,
            'stride_length_mean': 1.45,
            'stride_time_mean': 1.12,
            'walking_speed_mean': 1.29,
            'stance_phase_mean': 62.1
        },
        'joint_angles_101': {
            'hip_flexion_extension': [0] * 101,
            'knee_flexion_extension': [0] * 101,
            'ankle_dorsi_plantarflexion': [0] * 101
        },
        'analysis_timestamp': datetime.now().isoformat()
    }

    # Excel 형식 변환
    excel_path = "./output/S001_traditional_format.xlsx"
    Path("./output").mkdir(exist_ok=True)

    converter.convert_mediapipe_to_excel(sample_results, excel_path)

    # CSV 형식 변환
    csv_files = converter.convert_to_csv_format(sample_results, "./output/csv/")

    # 표준 JSON 형식 변환
    json_path = "./output/S001_standard_format.json"
    converter.convert_to_json_standard(sample_results, json_path)

    print("\n✅ 모든 형식 변환이 완료되었습니다!")
    print(f"📁 출력 파일:")
    print(f"  • Excel: {excel_path}")
    print(f"  • CSV: {csv_files}")
    print(f"  • JSON: {json_path}")

if __name__ == "__main__":
    main()